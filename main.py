from openpyxl import Workbook, load_workbook
from questions import orignal_column_headers, question_codes
from groupings import find_identifiers_for_group_name, groupings, group_identifiers, \
    find_relevant_questions_for_group_name

# superfluous cols coming out of limesurvey
columns_to_delete = [
    [190, 243],
    [9, 9],
    [4, 4]
]


def get_input_workbook():
    return load_workbook(filename="results-survey.xlsx")


def check_sheet_format():
    for column in processed_sheet.iter_cols(min_col=1, min_row=1, max_row=1):
        for cell in column:
            # remove duplicate whitespaces
            cell.value = ' '.join(cell.value.split())
            if cell.value.replace(u'\xa0', u' ').strip() != orignal_column_headers[cell.column].replace(u'\xa0',
                                                                                                        u' ').strip():
                print("ERROR: column does not have expected header")
                print("column id:", cell.column)
                print("column header", cell.value)
                print("expected header", orignal_column_headers[cell.column])


# many questions need to be reformated as they contain introduction texts, ...
def rename_headers():
    processed_sheet["K1"].value = "Randomization Group"

    if "Die kommenden Fragen beziehen sich" in processed_sheet["CG1"].value:
        processed_sheet["CG1"].value = "Wie einfach oder schwierig ist es für Sie, das Lastenrad unterwegs" \
                             " in Hamburg zu parken? (nicht an Ihrem Wohnort)"
    else:
        print("did not replace introduction: Die kommenden Fragen beziehen sich")

    if "Diese Frage bezieht sich auf das Parken von Lastenrädern im öffentlichen Raum" in processed_sheet["FL1"].value:
        processed_sheet["FL1"].value = "Wie einfach oder schwierig stellen Sie es vor," \
                             " ein Lastenrad unterwegs in Hamburg zu parken? (nicht am Wohnort)"
    else:
        print("did not replace introduction: Diese Frage bezieht sich auf das Parken von Lastenrädern im öffentlichen")

    if "Wir würden gerne auch soziodemografische Daten erheben. " in processed_sheet["FT1"].value:
        processed_sheet["FT1"].value = "Was ist Ihr Geschlecht?"
    else:
        print("did not replace introduction: Wir würden gerne auch soziodemografische Daten erheben.")

    if "Haben Sie vor sich ein eigenes Lastenrad anzuschaffen?" in processed_sheet["DJ1"].value:
        processed_sheet["DJ1"].value = "Haben Sie vor sich ein eigenes Lastenrad zu kaufen?"
    else:
        print("did not replace introduction: Haben Sie vor sich ein eigenes Lastenrad anzuschaffen?")

    # clean headers from double whitespaces etc.
    for cell in processed_sheet[1]:
        cell.value = ' '.join(cell.value.split())
        cell.value.replace(u'\xa0', u' ').strip()


def fix_empty_cells():
    for row in processed_sheet.iter_rows(min_row=2,
                                         min_col=1,
                                         ):
        for column in range(len(row)):
            if row[column].value == "N/A":
                row[column].value = None


# some questions are asked multiple times due to randomization groups.
# this function groups all results as a single question
def group_duplicate_questions(resulting_columns, duplicate_question_columns):
    # get the column headers for the resulting columns
    summary_columns_names = orignal_column_headers[resulting_columns[0]: (resulting_columns[1] + 1)]
    columns_to_delete.append(duplicate_question_columns)

    for row in processed_sheet.iter_rows(min_row=2, min_col=duplicate_question_columns[0], max_col=duplicate_question_columns[1]):
        for cell in row:
            if cell.value is not None:
                column_header = processed_sheet.cell(row=1, column=cell.column).value
                print(column_header)
                # try to find the current column header in the questions to be summarized.
                summary_column = summary_columns_names.index(column_header) + (
                resulting_columns[0])  # TODO: writing to the right cell?
                # copy value to summary columns
                processed_sheet.cell(row=cell.row, column=summary_column).value = cell.value


def export_question_times_to_separate_sheet():
    times_sheet = input_wb.copy_worksheet(processed_sheet)
    input_wb.active = times_sheet
    input_wb.active.title = "question_times"

    col_to_keep = [
        237,
        232,
        226,
        220,
        212,
        201,
        195,
        191,
        190,
        3,
        1
    ]

    # delete all columns that do not contain questions times
    delete_all_cols_but(times_sheet, col_to_keep)

    # set processed sheet back as active sheet
    input_wb.active = processed_sheet


# insert new header rows with question_ids
def add_question_and_subquestion_ids_to_cols():
    # add 2 rows for question_id and subquestion_id
    processed_sheet.insert_rows(idx=1, amount=2)
    for cell in processed_sheet[3]:
        header = cell.value
        for entry in question_codes:
            if entry[2] == header:
                # set question code
                processed_sheet.cell(row=1, column=cell.column).value = entry[0]
                # set subquestion code or None
                processed_sheet.cell(row=2, column=cell.column).value = entry[1]
                break
        else:
            # Didn't find anything..
            print("could not find question code for ", header)
            exit()


# create a new worksheet that lists an overview of all questions and subquestions
def export_question_ids_to_separate_sheet():
    question_codes_sheet = input_wb.copy_worksheet(input_wb.active)
    input_wb.active = question_codes_sheet
    input_wb.active.title = "question_codes"

    sheet = input_wb.active
    sheet.cell(row=1, column=1).value = "Question_IDs"
    sheet.cell(row=1, column=2).value = "Question_text"
    sheet.cell(row=1, column=3).value = "Sub_Question_ID"
    sheet.cell(row=1, column=4).value = "Sub_Question_text"

    listed_questions = []
    row_it = 2

    for entry in question_codes:
        if entry[0] not in listed_questions:
            listed_questions.append(entry[0])
            sheet.cell(row=row_it, column=1).value = entry[0]
            sheet.cell(row=row_it, column=2).value = entry[2].rsplit('[', 1)[0]
            sheet.cell(row=row_it, column=3).value = None
            sheet.cell(row=row_it, column=4).value = None
            row_it += 1
        else:
            sheet.cell(row=row_it, column=1).value = None
            sheet.cell(row=row_it, column=2).value = None
            sheet.cell(row=row_it, column=3).value = entry[1]
            sheet.cell(row=row_it, column=4).value = entry[2].rsplit('[', 1)[len(entry[2].rsplit('[', 1)) - 1]
            row_it += 1

    input_wb.active.delete_cols(5, amount=300)
    input_wb.active.delete_rows(row_it + 1, amount=300)

    # set processed sheet back as active sheet
    input_wb.active = processed_sheet


# create a seperate worksheet for each subgroup of survey-submissions
def export_group_subsets():
    for grouping in groupings:
        for group_name in grouping:
            if group_name in input_wb.sheetnames:
                continue

            # make group sheet
            group_sheet = input_wb.copy_worksheet(processed_sheet)
            input_wb.active = group_sheet
            input_wb.active.title = group_name
            rows_to_delete = []

            identifiers = find_identifiers_for_group_name(group_name)

            # loop through row and delete submissions with irrelevant data fro this group
            for row in group_sheet.iter_rows(min_row=4, min_col=1):
                for criteria in identifiers:
                    # check if criteria of identifier match for this row. otherwise delete row.
                    question_id = criteria[0]
                    relevant_answers = criteria[1]
                    question_col = find_col_of_question_id(group_sheet, 1, question_id)
                    current_cell_value = row[question_col - 1].value # row array starts counting at 0, col_id at 1
                    if current_cell_value not in relevant_answers:
                        row_id = row[0].row
                        rows_to_delete.append(row_id)

            # delete rows identified as irrelevant
            for row_to_del in reversed(rows_to_delete):
                group_sheet.delete_rows(idx=row_to_del)

            # loop through cols with question_ids and delete irrelevant cols
            cols_to_keep = []
            for question_id_cell in reversed(group_sheet[1]):
                # filter columns with irrelevant data
                if question_id_cell.value is not None:
                    for relevant_question in find_relevant_questions_for_group_name(group_name):
                        # question_id_cell.value can contain several question_ids
                        # as duplicated questions have been summarized
                        if relevant_question in question_id_cell.value:
                            cols_to_keep.append(question_id_cell.column)

            # delete cols identified as irrelevant
            delete_all_cols_but(group_sheet, cols_to_keep)

    input_wb.active = processed_sheet



def bulk_delete_columns(sheet, delete_those):
    # sort columns highest to lowest
    delete_those.sort(key=lambda x: x[0], reverse=True)

    for section in columns_to_delete:
        delete_from = section[0]
        section_length = section[1] - section[0] + 1
        sheet.delete_cols(idx=delete_from, amount=section_length)


def delete_all_cols_but(sheet, keep_these):
    for col in range(sheet.max_column + 1, 1, -1):
        if col not in keep_these:
            sheet.delete_cols(idx=col, amount=1)


def find_col_of_question_id(sheet, header_row, question_id):
    for cell in sheet[header_row]:
        if cell.value is not None and question_id in cell.value:
            return cell.column

    return None


# cleans dataset by deleting superfluous columns, summarizing duplicate questions
# exports sub-datasets to separate worksheets
if __name__ == '__main__':
    input_wb = get_input_workbook()
    input_wb.active.title = "raw_data"

    processed_sheet = input_wb.copy_worksheet(input_wb.active)
    input_wb.active = processed_sheet
    processed_sheet.title = "processed_data"

    check_sheet_format()

    fix_empty_cells()

    rename_headers()

    # summarize challenges
    group_duplicate_questions([25, 34], [35, 74])

    # summarize improvements
    group_duplicate_questions([75, 76], [77, 84])

    # summarize criteria
    group_duplicate_questions([96, 100], [101, 105])

    # summarize criteria (non-user)
    group_duplicate_questions([96, 100], [170, 174])

    # summarize imagined challenges [non-user]
    group_duplicate_questions([118, 127], [128, 167])

    # summarize willingness to purchase and hurdles in purchase
    group_duplicate_questions([22, 24], [114, 116])

    # summarize "Kennen Sie den lastenradstellplatz"
    group_duplicate_questions([95, 95], [169, 169])

    # summarize "Haben Sie Vorschläge..."
    group_duplicate_questions([107, 107], [175, 175])

    export_question_times_to_separate_sheet()

    export_question_ids_to_separate_sheet()

    # delete all columns that became redundant during summaries
    bulk_delete_columns(processed_sheet, columns_to_delete)

    # insert new header rows with question_ids
    add_question_and_subquestion_ids_to_cols()

    export_group_subsets()

    # TODO Test if all cells are filled in that should be filled

    input_wb.save('cargo_bike_survey_answers.xlsx')
    print("finished!")
